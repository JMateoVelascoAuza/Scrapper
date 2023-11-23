import openpyxl
import time
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup
import tkinter as tk


def get_urlLinkedin_and_name(names, last_name,driver):
    upb = url_entry.get()
    for name in names:
        try:
            driver.get(upb + '?keywords=' + name + ' ' + last_name)
            
            time.sleep(3)
            src = driver.page_source

            soup = BeautifulSoup(src, 'html.parser')
            
            pav = soup.find('section', {'class' : 'artdeco-card artdeco-card--with-hover ember-view full-width org-people-profile-card'})
            all_links = pav.find('a', {'class' : 'app-aware-link'})
            user_name_container = pav.find('div',{'class':'ember-view lt-line-clamp lt-line-clamp--single-line org-people-profile-card__profile-title t-black'}).get_text()
            linkedin = all_links.get('href')
            user_name = user_name_container.replace('\n','').replace('\t','')
            print(linkedin,user_name)
            return [linkedin.split('?')[0],user_name]       
        except:
            pass
    
    return 'N/A'

def update_spreadsheet():
    # get the input values
    email = email_entry.get()
    password = password_entry.get()
    excel_location = excel_location_entry.get()
    sheet_name = sheet_name_entry.get()

    # open the Excel workbook and get the sheet
    wb = openpyxl.load_workbook(excel_location)
    sheet = wb[sheet_name]

    # create a list of names and a list of last names
    names = []
    last_names = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        names.append(row[2])
        last_names.append(row[3])

    # open the LinkedIn website
    driver = webdriver.Chrome('C:/Users/UPB/Desktop/Scrapper/buscarurl/chromedriver.exe')
    driver.get("https://linkedin.com/uas/login")

    # log in to LinkedIn
    username = driver.find_element("id", "username")
    password_field = driver.find_element('id','password')

    username.send_keys(email)
    password_field.send_keys(password)

    password_field.send_keys(Keys.RETURN)

    time.sleep(20)

    # go to the alumni search page
    upb = url_entry.get()
    driver.get(upb)

    linkedins = []

    # loop through the names and last names
    for name, last_name in zip(names, last_names):
        # get the LinkedIn url
        linkedin = get_urlLinkedin_and_name(str(name).split(),str(last_name),driver)
        linkedins.append(linkedin)
        print(repr(linkedin))
    
    # close the browser
    driver.quit()

    # Write the values to the spreadsheet, starting at cell A1
    for i, value in enumerate(linkedins,start=2):
        for j,user_data in enumerate(value):
            sheet.cell(row=i+1, column=7+j).value = user_data

    # Save the changes to the file
    wb.save(excel_location)

    # show a message box with a success message
    # tk.messagebox.showinfo("Success", "The LinkedIn URLs have been updated successfully!")
    driver.quit()

# create the main window
root = tk.Tk()

# set the title
root.title("LinkedIn URL Updater")

# create the email label and entry
email_label = tk.Label(root, text="Email")
email_label.grid(row=0, column=0)

email_entry = tk.Entry(root)
email_entry.grid(row=0, column=1)

# create the password label and entry
password_label = tk.Label(root, text="Password")
password_label.grid(row=1, column=0)

password_entry = tk.Entry(root, show="*")
password_entry.grid(row=1, column=1)

# create the Excel location label and entry
excel_location_label = tk.Label(root, text="Excel Location")
excel_location_label.grid(row=3, column=0)

excel_location_entry = tk.Entry(root)
excel_location_entry.grid(row=3, column=1)

##create the sheet name label and entry
sheet_name_label = tk.Label(root, text="Sheet Name")
sheet_name_label.grid(row=4, column=0)

sheet_name_entry = tk.Entry(root)
sheet_name_entry.grid(row=4, column=1)

##create the URL label and entry
url_label = tk.Label(root, text="LinkedIn Alumni URL")
url_label.grid(row=5, column=0)

url_entry = tk.Entry(root)
url_entry.grid(row=5, column=1)

##create the update button
update_button = tk.Button(root, text="Update LinkedIn URLs", command=update_spreadsheet)
update_button.grid(row=6, column=0, columnspan=2)

##run the main loop
root.mainloop()




