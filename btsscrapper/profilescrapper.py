import time
from bs4 import BeautifulSoup
from selenium import webdriver
import openpyxl
import PySimpleGUI as sg
from selenium.webdriver.chrome.options import Options

def extract_section(sections, name_section):
    for section in sections:
        if section.find('div', {'id': name_section}) is not None:
            experience = section
            break
    try:
        all_works = experience.find_all(
            'div', {'class': 'display-flex flex-column full-width align-self-center'})
        info_experience = ''
        for work in all_works:
            for infowork in work.find_all('span', {'class': 'visually-hidden'}):
                info_experience += f'{infowork.get_text().strip()}\n'
            info_experience += '~\n'
    except:
        info_experience = 'N/A'

    return info_experience

def extract_profile_info(linkedin_url, driver):
    user_info = {
        'name': 'N/A',
        'headline': 'N/A',
        'workplace': 'N/A',
        'city': 'N/A',
        'university': 'N/A',
        'is_open': 'N/A',
        'info_experience': 'N/A',
        'info_education': 'N/A',
        'info_skills': 'N/A'
    }

    try:
        driver.get(linkedin_url)
        time.sleep(10)
        soup = BeautifulSoup(driver.page_source, 'html.parser')
        header = soup.find('section', {'class': 'artdeco-card ember-view pv-top-card'})
        sections = soup.find_all('section', {'class': 'artdeco-card ember-view relative break-words pb3 mt2'})

        try:
            user_info['name'] = header.find('h1', {'class': 'text-heading-xlarge inline t-24 v-align-middle break-words'}).get_text().strip()
        except Exception as e:
            print(f"Error extracting name: {e}")

        try:
            user_info['headline'] = header.find('div', {'class': 'text-body-medium break-words'}).get_text().strip()
        except Exception as e:
            print(f"Error extracting headline: {e}")

        try:
            user_info['workplace'] = header.find('span', {'class': 'pv-text-details__right-panel-item-text hoverable-link-text break-words text-body-small t-black'}).get_text().strip()
        except Exception as e:
            print(f"Error extracting workplace: {e}")

        try:
            user_info['city'] = header.find('span', {'class': 'text-body-small inline t-black--light break-words'}).get_text().strip()
        except Exception as e:
            print(f"Error extracting city: {e}")

        try:
            university = header.find('span', {'class': 'text-body-small inline t-black--light break-words'})
            user_info['university'] = university.get_text().strip() if university else 'N/A'
        except Exception as e:
            print(f"Error extracting university: {e}")

        try:
            is_open_text = header.find('h3', {'class': 'truncate text-body-small'})
            user_info['is_open'] = 'Open to work' if 'Open to work' in (is_open_text.get_text().strip() if is_open_text else '') else 'Not open to work'
        except Exception as e:
            print(f"Error extracting is_open: {e}")

        try:
            user_info['info_experience'] = extract_section(sections, 'experience')
        except Exception as e:
            print(f"Error extracting info_experience: {e}")

        try:
            user_info['info_education'] = extract_section(sections, 'education')
        except Exception as e:
            print(f"Error extracting info_education: {e}")

        try:
            user_info['info_skills'] = extract_section(sections, 'skills')
        except Exception as e:
            print(f"Error extracting info_skills: {e}")
        
        print(user_info)
        return user_info

    except Exception as e:
        print(f"Error extracting profile info: {e}")
        return user_info

    


def main():
    # Create the PySimpleGUI layout for the login credentials window
    layout = [
        [sg.Text('LinkedIn Login')],
        [sg.Text('Email'), sg.Input(key='email')],
        [sg.Text('Password'), sg.Input(key='password', password_char='*')],
        [sg.Text('Excel File Name'), sg.Input(key='filename')],
        [sg.Text('Sheet Name'), sg.Input(key='sheetname')],
        [sg.Button('Login')]
    ]

    # Create the PySimpleGUI window
    window = sg.Window('Login to LinkedIn', layout)

    # Wait for the user to enter their login credentials and Excel file name
    event, values = window.read()

    # Create a webdriver instance
    driver = webdriver.Edge('msedgedriver.exe')
    # Navigate to the LinkedIn login page
    driver.get("https://linkedin.com/uas/login")

    # Enter login credentials
    email = driver.find_element("id", "username")
    email.send_keys(values['email'])

    password = driver.find_element('id', 'password')
    password.send_keys(values['password'])

    # Click login button
    driver.find_element('xpath', "//button[@type='submit']").click()

    # Wait for the page to load
    time.sleep(20)

    data_linkedin = []

    # Load the Excel file
    wb = openpyxl.load_workbook(f"{values['filename']}.xlsx".replace('\\','/'))
    sheet = wb[values['sheetname']]

    # Create a list of LinkedIn profile URLs to scrape
    links = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        links.append(row[0])

    # Scrape data from each LinkedIn profile
    for linkedin_url in links:
        print(linkedin_url)
        userinfo = extract_profile_info(linkedin_url,driver)

        data_linkedin.append(userinfo)

    # Close the webdriver and the PySimpleGUI window
    driver.close()
    window.close()

    # Write the scraped data to the Excel file
    for i, userinfo2 in enumerate(data_linkedin, start=2):  # Start at row 2
        for j, key in enumerate(userinfo2.keys(), start=2):  # Start at column 1
            # Write the header only once (for the first row)
            if i == 2:
                sheet.cell(row=1, column=j).value = key
            # Write the user info
            sheet.cell(row=i, column=j).value = userinfo2[key]

    wb.save(f"{values['filename']}OUT.xlsx")


main()
